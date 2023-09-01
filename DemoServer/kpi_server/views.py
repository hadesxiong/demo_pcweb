from django.shortcuts import render

from django.core.paginator import Paginator
from django.db.models import Subquery,Q
from django.db import IntegrityError

from django.http.response import JsonResponse
from rest_framework.decorators import api_view

from kpi_server.models import Users,Org,Reference

from kpi_server.serializers import UsersSerializer,OrgSerializer,RefSerializer

import math,datetime

from openpyxl import load_workbook
# Create your views here.

'''查询用户列表'''

@api_view(['GET'])
def getUsersList(request):

    # params解析
    query_params = {
        'belong_group': int(request.query_params.get('group',0)),
        'user_character': int(request.query_params.get('character',0)),
        'org_level': int(request.query_params.get('org',0)),
        'key_word':request.query_params.get('ext',None),
        'user_client': request.query_params.get('client',None),
        'page_size':int(request.query_params.get('size',15)),
        'page':int(request.query_params.get('page',1))
    }
    
    # 筛选出机构层级
    if query_params['org_level'] == 0:
        org_query = Org.objects.all().values('org_num','org_name')
    else:
        org_query = Org.objects.filter(org_level=query_params['org_level']).values('org_num','org_name')

    # 关键词
    if query_params['key_word'] is not None:
        complex_condition = Q(notes_id__contains=query_params['key_word']) | Q(user_name__contains=query_params['key_word']) | Q(user_belong_org__in=Subquery(Org.objects.filter(org_name__contains=query_params['key_word']).values('org_num')))
    else:
        complex_condition = Q()

    # 筛选用户所属条线
    if query_params['belong_group'] == 0:
        group_condition = Q()        
    else:
        group_condition = Q(user_belong_group=query_params['belong_group'])

    # 筛选用户角色
    if query_params['user_character'] == 0:
        character_condition = Q()
    else:
        character_condition = Q(user_character=query_params['user_character'])

    users_querySet = Users.objects.filter(user_belong_org__in=Subquery(org_query.values('org_num'))).filter(group_condition).filter(character_condition).filter(complex_condition)

    # 分页
    page_inator = Paginator(users_querySet,query_params['page_size'])
    page_max = math.ceil(len(users_querySet)/query_params['page_size'])

    if query_params['page'] <= page_max:
        each_users_list = page_inator.page(query_params['page'])
        users_serializer = UsersSerializer(each_users_list,many=True)
        re_msg = {'data':users_serializer.data,'code':0,'msg':'success','has_next':(query_params['page']<page_max)}
    else:
        re_msg = {'code':1,'msg':'error_range'}

    return JsonResponse(re_msg,safe=False)

'''删除/更新用户信息'''

@api_view(['POST'])
def updateUser(request):

    # body解析
    body_data = {
        'type':request.data.get('type','update'),
        'notes_id':request.data.get('user',None),
        'update_data':request.data.get('update_data',None),
    }

    if body_data['notes_id'] is not None:
        user_obj = Users.objects.get(notes_id=body_data['notes_id'])

        if user_obj and  body_data['update_data'] is not None and (body_data['type'] == 'update'):
            for (x,y) in body_data['update_data'].items():
                print(x,y)
                setattr(user_obj,x,y)
            user_obj.save()
            re_msg = {'code':0,'msg':'update success'}

        elif user_obj and (body_data['type'] == 'delete'):
            user_obj.delete()
            re_msg = {'code':0,'msg':'delete success'}

        else:
            re_msg = {'code':1,'msg':'user not found'}

    else:
        re_msg = {'code':3,'msg':'error params.'}

    return JsonResponse(re_msg,safe=False)

'''新增用户'''
@api_view(['POST'])
def createUser(request):

    # body解析
    body_data = {
        'is_multi':request.data.get('is_multi',False),
        'update_data':request.data.get('update_data',None),
        'update_file':request.FILES.get('update_file',None)
    }

    print(body_data)

    # 查询提交的内容
    if body_data['is_multi'] and body_data['update_file'] is not None:

        # excel文件校验
        # 解析文件
        xls_files = load_workbook(body_data['update_file'])
        xls_sheet = xls_files['用户导入模板']

        user_list = []
        for row in xls_sheet.iter_rows(min_row=7,values_only=True):
            each_user = {
                'notes_id' : row[0],
                'user_name': row[1],
                'user_character':int(row[2].split('-')[0]),
                'user_belong_group':int(row[3].split('-')[0]),
                'user_belong_org':int(row[4]),
                'user_state':0,
                'user_create': datetime.date.today().strftime("%Y-%m-%d"),
                'user_update': datetime.date.today().strftime("%Y-%m-%d"),
            }
            user_list.append(each_user)

        # try:
        #     # 批量添加
        #     Users.objects.bulk_create(user_list)

        # except IntegrityError as e:
        #     # 唯一约束错误
        #     print('唯一约束错误',e)

        err_row_list = []

        for index,each_user in enumerate(user_list):
            obj,created = Users.objects.get_or_create(**each_user)
            # 插入失败的存入err_row_list
            if created:
                pass
            else:
                err_row_list.append(index)

        re_msg = {'code':1,'err':err_row_list,'msg':'done'}

    else:
        # 查询notesid是否存在，

        try:
            user_obj = Users.objects.get(notes_id=body_data.get('update_data').get('notes_id'))
            re_msg = {'code':'2','msg':'users exists.'}

        except Users.MultipleObjectsReturned:
            re_msg = {'code':3,'msg':'multi users exists'}

        except Users.DoesNotExist:
            new_user = Users()
            for (x,y) in body_data.get('update_data').items():
                setattr(new_user,x,y)
            new_user.save()
            re_msg = {'code':0,'msg':'create_success'}

    return JsonResponse(re_msg,safe=False)

'''查询机构列表'''
@api_view(['GET'])
def getOrgList(request):

    # params解析
    query_params = {
        'org_level': int(request.query_params.get('level',0)),
        'org_group': int(request.query_params.get('group',0)),
        'key_word': request.query_params.get('ext',None),
        'user_client': request.query_params.get('client',None),
        'page_size':int(request.query_params.get('size',15)),
        'page':int(request.query_params.get('page',1))
    }

    # 分组筛选条件
    if query_params['org_group'] == 0:
        group_condition = Q()
    else:
        group_condition = Q(org_group=query_params['org_group'])

    # 层级筛选条件
    if query_params['org_level'] == 0:
        level_condition = Q()
    else:
        level_condition = Q(org_level=query_params['org_level'])

    # 关键词筛选
    if query_params['key_word'] is not None:
        complex_condition = Q(org_name__contains=query_params['key_word']) | Q(org_num__contains=query_params['key_word']) | Q(org_manager__in=Subquery(Users.objects.filter(notes_id__contains=query_params['key_word']).values('notes_id'))) | Q(org_manager__in=Subquery(Users.objects.filter(notes_id__contains=query_params['key_word']).values('notes_id')))
    else:
        complex_condition = Q()

    org_querySet = Org.objects.filter(group_condition).filter(level_condition).filter(complex_condition)

    # 分页
    page_inator = Paginator(org_querySet,query_params['page_size'])
    page_max = math.ceil(len(org_querySet)/query_params['page_size'])

    if query_params['page'] <= page_max:
        each_orgs_list = page_inator.page(query_params['page'])
        orgs_serializer = OrgSerializer(each_orgs_list,many=True)
        re_msg = {'data':orgs_serializer.data,'code':0,'msg':'success','has_next':(query_params['page']<page_max)}
    else:
        re_msg = {'code':1,'msg':'error_range'}

    return JsonResponse(re_msg,safe=False)

'''新增/更新/删除机构'''
@api_view(['POST'])
def updateOrg(request):

    body_data = {
        'update_type': request.data.get('type','create'),
        'org_num': request.data.get('num',None),
        'update_data': request.data.get('update_data',None)
    }

    # 新增逻辑
    if body_data['update_type'] == 'create':
        # 表单
        update_data = {
            'org_num':body_data['org_num'],
            'org_name':body_data['update_data']['org_name'],
            'parent_org_id':body_data['update_data']['parent_org'],
            'org_level':body_data['update_data']['org_level'],
            'org_group':body_data['update_data']['org_group'],
            'org_manager':body_data['update_data']['org_manager'],
            'org_state':0,
            'org_create':datetime.date.today().strftime("%Y-%m-%d"),
            'org_update':datetime.date.today().strftime("%Y-%m-%d")
        }
        # 检验是否为空
        if None in update_data.values():
            re_msg = {'code':1,'msg':'err params'}

        else:
            obj,created = Org.objects.get_or_create(**update_data)
            if created:
                re_msg = {'code':0,'msg':'created','err':''}
            else:
                re_msg = {'code':1,'msg':'exists','err':''}

    # 更新逻辑
    elif body_data['update_type'] == 'update' and body_data['update_data'] is not None:

        try:
            org_obj = Org.objects.get(org_num=body_data['org_num'])
            for (x,y) in body_data['update_data'].items():
                setattr(org_obj,x,y)
            
            org_obj.save()
            re_msg = {'code':0,'msg':'updated','err':''}
        
        except Org.DoesNotExist:
            re_msg = {'code':1,'msg':'org not exists'}

    elif body_data['update_type'] == 'delete':

        try:
            org_obj = Org.objects.get(org_num=body_data['org_num'])
            org_obj.delete()

            re_msg = {'code':0,'msg':'deleted','err':''}
        
        except Org.DoesNotExist:
            re_msg = {'code':1,'msg':'org not exists'}

    return JsonResponse(re_msg,safe=False)

'''筛选项处理-处筛选'''
@api_view(['GET'])
def getFilter(request):

    # params解析
    query_params = {
        'ref_type': request.query_params.get('type',None),
        'ref_scene': int(request.query_params.get('scene',0)),
        'client': request.query_params.get('client','client')
    }

    if None in query_params.values():
        re_msg = {'code':1,'msg':'error params'}
    else:
        # 拆分ref_type
        type_list = query_params['ref_type'].split('.')
        ref_queryset = Reference.objects.filter(ref_ext__in=type_list)

        ref_serializer = RefSerializer(ref_queryset,many=True)

        re_msg = {'code':0,'data':ref_serializer.data}

    return JsonResponse(re_msg,safe=False)
